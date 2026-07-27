import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ═══ Styles ═══
NAVY = "0F172A"
BLUE = "1D4ED8"
GREEN = "059669"
AMBER = "D97706"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "D1FAE5"
LIGHT_YELLOW = "FEF3C7"
LIGHT_PURPLE = "EDE9FE"
MID_GRAY = "D1D5DB"

header_font = Font(bold=True, color=WHITE, size=11, name="Arial")
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
meta_label_font = Font(bold=True, size=10, name="Arial")
meta_value_font = Font(size=10, name="Arial")
section_font = Font(bold=True, color=WHITE, size=12, name="Arial")
hint_font = Font(italic=True, color="9CA3AF", size=9, name="Arial")
value_font = Font(size=10, name="Arial")
label_font = Font(bold=True, size=10, name="Arial")
thin_border = Border(
    left=Side(style="thin", color=MID_GRAY),
    right=Side(style="thin", color=MID_GRAY),
    top=Side(style="thin", color=MID_GRAY),
    bottom=Side(style="thin", color=MID_GRAY),
)
wrap_top = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")

category_fills = {
    "T2 Issue": LIGHT_BLUE,
    "Tracker": LIGHT_PURPLE,
    "Customer Conversation": LIGHT_GREEN,
    "Email Conversation": LIGHT_YELLOW,
    "Appendix": LIGHT_GRAY,
}

def styled(ws, row, col, value=None, font=value_font, fill=None, align=wrap_top):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    return cell

def section_bar(ws, row, text, color, cols=11):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = section_font
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, cols + 1):
        ws.cell(row=row, column=c).border = thin_border
    ws.row_dimensions[row].height = 30

def build_month_sheet(ws, month, year, author, pub_month, pub_year, tickets_data, counts):
    """Build a single month sheet with meta + tickets combined."""
    COLS = 11

    # ── Row 1: Title bar ──
    section_bar(ws, 1, f"  FAE Support Report — {month} {year}", BLUE, COLS)
    ws.row_dimensions[1].height = 36

    # ── Rows 2-5: Meta info (2-column pairs) ──
    meta_layout = [
        # row 2
        [("Report Month", month, 1), ("Author", author, 5), ("Total Tickets", str(counts["total"]), 9)],
        # row 3
        [("Report Year", str(year), 1), ("Published", f"{pub_month} {pub_year}", 5), ("Resolved", str(counts["resolved"]), 9)],
        # row 4
        [("T2 Issue", str(counts["t2"]), 1), ("Tracker", str(counts["tracker"]), 3),
         ("Customer Conv.", str(counts["conv"]), 5), ("Email Conv.", str(counts["email"]), 7),
         ("Open", str(counts["open"]), 9), ("Carry-over", str(counts["carryover"]), 11)],
    ]

    # Row 2
    for label, val, col in meta_layout[0]:
        styled(ws, 2, col, label, font=meta_label_font, fill=LIGHT_GRAY)
        styled(ws, 2, col + 1, val, font=meta_value_font)
    # Row 3
    for label, val, col in meta_layout[1]:
        styled(ws, 3, col, label, font=meta_label_font, fill=LIGHT_GRAY)
        styled(ws, 3, col + 1, val, font=meta_value_font)
    # Row 4: compact stats
    for label, val, col in meta_layout[2]:
        styled(ws, 4, col, label, font=meta_label_font, fill=LIGHT_GRAY)
        if col + 1 <= COLS:
            styled(ws, 4, col + 1, val, font=meta_value_font)

    # ── Row 5: empty spacer ──
    ws.row_dimensions[5].height = 8

    # ── Row 6: Ticket section header ──
    section_bar(ws, 6, "  Ticket Data", GREEN, COLS)

    # ── Row 7: Column headers ──
    ticket_headers = [
        "Category", "Ticket ID", "Title", "Affected Product/Model",
        "Open Date", "Last Response", "Close Date", "Status",
        "Case Description", "Case Summary", "Ticket Details",
    ]
    for i, h in enumerate(ticket_headers, 1):
        styled(ws, 7, i, h, font=header_font, fill=NAVY, align=Alignment(horizontal="center", vertical="center", wrap_text=True))
    ws.row_dimensions[7].height = 30

    # ── Row 8: Hints ──
    hints = [
        "T2 Issue / Tracker / Customer Conversation / Email Conversation / Appendix",
        "e.g., 82514414", "Short title", "e.g., KOKOMO-W220AX-IS",
        "MMM DD, YYYY", "MMM DD, YYYY", "MMM DD, YYYY", "Resolved / Open",
        "What the customer reported", "Use • for bullet points", "Full narrative (optional for Conversation)",
    ]
    for i, hint in enumerate(hints, 1):
        styled(ws, 8, i, hint, font=hint_font, fill=LIGHT_GRAY)
    ws.row_dimensions[8].height = 40

    # ── Row 9+: Ticket data ──
    for row_idx, ticket in enumerate(tickets_data, 9):
        cat = ticket[0]
        fill = category_fills.get(cat, None)
        for col_idx, val in enumerate(ticket, 1):
            styled(ws, row_idx, col_idx, val, fill=fill)
        ws.row_dimensions[row_idx].height = 100

    # ── Column widths ──
    col_widths = [22, 14, 35, 24, 14, 14, 14, 12, 45, 55, 55]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze below headers ──
    ws.freeze_panes = "A9"

    # ── Data validations ──
    cat_dv = DataValidation(
        type="list",
        formula1='"T2 Issue,Tracker,Customer Conversation,Email Conversation,Appendix"',
    )
    ws.add_data_validation(cat_dv)
    cat_dv.add("A9:A100")

    status_dv = DataValidation(type="list", formula1='"Resolved,Open"')
    ws.add_data_validation(status_dv)
    status_dv.add("H9:H100")


# ═══════════════════════════════════════
# Sheet 1: Instructions (fixed)
# ═══════════════════════════════════════
ws_inst = wb.active
ws_inst.title = "Instructions"
ws_inst.sheet_properties.tabColor = AMBER

ws_inst.merge_cells("A1:B1")
t = ws_inst["A1"]
t.value = "How to Use This Template"
t.font = Font(bold=True, color=WHITE, size=14, name="Arial")
t.fill = PatternFill(start_color=AMBER, end_color=AMBER, fill_type="solid")
t.alignment = center
ws_inst.row_dimensions[1].height = 36

instructions = [
    ("Structure", "Each month gets ONE tab. The tab contains report info at the top and ticket data below."),
    ("", ""),
    ("Monthly", "1. Duplicate an existing month tab (right-click → Duplicate)"),
    ("Workflow", "2. Rename the new tab to YYYY-MM (e.g., 2026-04)"),
    ("", "3. Update the report info rows at the top (month, counts, etc.)"),
    ("", "4. Clear the old ticket rows and fill in new tickets"),
    ("", "5. Share the file with Terrel or Claude"),
    ("", ""),
    ("Category", "• T2 Issue / Tracker → fill all 3: Description, Summary, Details"),
    ("Guide", "• Customer Conversation / Email Conversation → Description = Highlight, Summary = Summary, Details = leave empty"),
    ("", "• Appendix → Description = Highlight only"),
    ("", ""),
    ("Formatting", "• Case Summary: use • at the start of each bullet point"),
    ("Tips", "• Affected Product/Model: tag the specific device (KOKOMO-W220AX-IS, ESG620, etc.)"),
    ("", "• Category and Status columns have dropdowns"),
    ("", "• Newest month tab should be leftmost (after Instructions)"),
]

for idx, (label, text) in enumerate(instructions, 3):
    cell_a = ws_inst.cell(row=idx, column=1, value=label)
    cell_a.font = label_font if label else value_font
    cell_b = ws_inst.cell(row=idx, column=2, value=text)
    cell_b.font = value_font
    cell_b.alignment = wrap_top

ws_inst.column_dimensions["A"].width = 14
ws_inst.column_dimensions["B"].width = 85


# ═══════════════════════════════════════
# Sheet 2: 2026-03 (example with data)
# ═══════════════════════════════════════
ws_mar = wb.create_sheet("2026-03", 1)  # insert after Instructions
ws_mar.sheet_properties.tabColor = BLUE

march_tickets = [
    [
        "T2 Issue", "82514414",
        "Summary of RADIUS Test Failure (KOKOMOCloud / APRESIA)",
        "KOKOMO-W220AX-IS",
        "Mar 12, 2026", "Mar 30, 2026", "Apr 2, 2026", "Resolved",
        "When testing the KOKOMOCloud feature while using Custom RADIUS under WPA3-Enterprise, a certificate missing error is displayed even though RadSec is disabled. The same error is observed from APRESIA.",
        "• The Phenomenon: RADIUS test under WPA3-Enterprise showed certificate missing error even with RadSec disabled\n• Root Cause (AP Defect): Protocol mismatch — script hardcoded GTC, customer required MS-CHAPv2\n• Misleading Error Logic: Generic RadSec error message for any RADIUS failure\n• Immediate Resolution: R&D patched AP via SSH to support both MS-CHAPv2 and GTC\n• Long-term: Fix in next firmware release + UI error message update (ETA: Late April 2026)",
        "In March 2026, a significant technical discrepancy was identified within the KOKOMOCloud and APRESIA monitoring platforms concerning the KOKOMO-W220AX-IS Access Point (AP). The primary issue involved a persistent 'certificate missing' error displayed during the RADIUS Test function, specifically when utilizing WPA3-Enterprise security...",
    ],
    [
        "T2 Issue", "83633969",
        '"Web Authentication Failed" Error',
        "KOKOMOCloud (Web Portal)",
        "Mar 16, 2026", "Mar 19, 2026", "Apr 2, 2026", "Resolved",
        "User so-aoi@iij.ad.jp reported unable to log in from March 11-16 due to 'web authentication failed' error. Issue self-resolved. Attempted different browsers, private windows, password reset — none worked.",
        "• Root Cause (Google reCAPTCHA): Error mapped to Google reCAPTCHA v3 — login blocked due to low risk score\n• Risk Factors: Low IP reputation, VPN/proxy, ad blockers, mechanical browsing patterns\n• Persistent Failure: reCAPTCHA evaluates IP-level, switching browsers didn't help\n• Resolution over Time: Resolved naturally as IP reputation improved\n• Distinction from Credential Errors: 'web auth failed' occurs before 2FA — confirms security gate failure\n• Recommended Mitigation: Switch networks (mobile hotspot) or implement whitelist mechanism",
        "Between March 11 and March 16, 2026, a persistent login issue was reported for the user account so-aoi@iij.ad.jp...",
    ],
    [
        "T2 Issue", "86120230",
        "Unable to Use Activity and Internet Connectivity in Diagnostic Tools",
        "KOKOMOCloud (Diagnostic Tools)",
        "Mar 26, 2026", "Mar 31, 2026", "Mar 31, 2026", "Resolved",
        "Activity and Internet Connectivity pages in diagnostic tools fail to load — stuck on loading spinner. Only occurs in corporate network with TLS/SSL inspection proxy. Other tools (Cable Test, Traceroute) work fine.",
        "• Issue: Diagnostic pages stuck on continuous loading screen\n• Environment: Corporate network with TLS/SSL inspection proxy only\n• Root Cause: Pages rely on WebSocket (WSS over port 443) — proxy disrupts handshake\n• Resolution: Exclude dolphin-jp-*.production.kokomocloud.com from TLS/SSL inspection\n• Outcome: Customer guidance provided, ticket closed",
        "This report outlines an issue encountered when accessing the Activity and Internet Connectivity pages within the diagnostic tools of a cloud-managed device...",
    ],
]

build_month_sheet(
    ws_mar, "March", 2026, "Himadri Paul, Senao Network",
    "April", "2026", march_tickets,
    {"t2": 3, "tracker": 0, "conv": 0, "email": 0, "total": 3, "resolved": 3, "open": 0, "carryover": 0},
)


# ═══════════════════════════════════════
# Sheet 3: 2026-04 (blank template)
# ═══════════════════════════════════════
ws_apr = wb.create_sheet("2026-04", 2)
ws_apr.sheet_properties.tabColor = GREEN

build_month_sheet(
    ws_apr, "April", 2026, "Himadri Paul, Senao Network",
    "May", "2026", [],
    {"t2": 0, "tracker": 0, "conv": 0, "email": 0, "total": 0, "resolved": 0, "open": 0, "carryover": 0},
)


# ═══ Save ═══
output = "/Users/terrelyeh/Downloads/Temp/FAE Report/FAE_Report_Template_v2.xlsx"
wb.save(output)
print(f"Saved: {output}")
