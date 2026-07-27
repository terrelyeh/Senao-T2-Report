import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ═══ Colors & Styles ═══
NAVY = "0F172A"
BLUE = "1D4ED8"
WHITE = "FFFFFF"
LIGHT_GRAY = "F3F4F6"
LIGHT_BLUE = "DBEAFE"
LIGHT_GREEN = "D1FAE5"
LIGHT_YELLOW = "FEF3C7"
LIGHT_PURPLE = "EDE9FE"

header_font = Font(bold=True, color=WHITE, size=11, name="Arial")
header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
section_font = Font(bold=True, color=BLUE, size=11, name="Arial")
section_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
label_font = Font(bold=True, size=10, name="Arial")
value_font = Font(size=10, name="Arial")
hint_font = Font(italic=True, color="9CA3AF", size=9, name="Arial")
thin_border = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
wrap_align = Alignment(wrap_text=True, vertical="top")

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_cell(ws, row, col, font=value_font, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.font = font
    cell.alignment = wrap_align
    cell.border = thin_border
    if fill:
        cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    return cell


# ═══════════════════════════════════════
# Sheet 1: Report Meta
# ═══════════════════════════════════════
ws1 = wb.active
ws1.title = "Report Info"
ws1.sheet_properties.tabColor = BLUE

meta_fields = [
    ("Report Month", "March", "e.g., January, February, March..."),
    ("Report Year", "2026", ""),
    ("Author", "Himadri Paul, Senao Network", "Report author name"),
    ("Publish Month", "April", "Month when report is published"),
    ("Publish Year", "2026", ""),
    ("T2 Issue Tickets", "3", "Number of T2 Issue Tickets"),
    ("Tracker Tickets", "0", "Number of Tracker Tickets"),
    ("Customer Conversation", "0", "Number of Customer Conversations"),
    ("Email Conversation", "0", "Number of Email Conversations"),
    ("Total Tickets", "3", "Sum of all ticket categories"),
    ("Resolved Tickets", "3", ""),
    ("Open Tickets", "0", ""),
    ("Carry-over from Previous", "0", "Tickets from previous month still open"),
]

# Title
ws1.merge_cells("A1:C1")
title_cell = ws1["A1"]
title_cell.value = "FAE Support Report — Monthly Info"
title_cell.font = Font(bold=True, color=WHITE, size=14, name="Arial")
title_cell.fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 36

# Headers
headers = ["Field", "Value", "Notes"]
for i, h in enumerate(headers, 1):
    cell = ws1.cell(row=2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

for idx, (field, value, note) in enumerate(meta_fields, 3):
    style_cell(ws1, idx, 1, font=label_font).value = field
    style_cell(ws1, idx, 2).value = value
    style_cell(ws1, idx, 3, font=hint_font).value = note

ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 35
ws1.column_dimensions["C"].width = 40


# ═══════════════════════════════════════
# Sheet 2: Tickets
# ═══════════════════════════════════════
ws2 = wb.create_sheet("Tickets")
ws2.sheet_properties.tabColor = "059669"

ticket_headers = [
    "Category",
    "Ticket ID",
    "Title",
    "Affected Product/Model",
    "Open Date",
    "Last Response",
    "Close Date",
    "Status",
    "Case Description",
    "Case Summary",
    "Ticket Details",
]

# Title
ws2.merge_cells("A1:K1")
t2 = ws2["A1"]
t2.value = "FAE Support Report — Ticket Data"
t2.font = Font(bold=True, color=WHITE, size=14, name="Arial")
t2.fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 36

# Headers
for i, h in enumerate(ticket_headers, 1):
    cell = ws2.cell(row=2, column=i, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border
ws2.row_dimensions[2].height = 30

# Hint row
hints = [
    "T2 Issue / Tracker / Customer Conversation / Email Conversation / Appendix",
    "e.g., 82514414",
    "Short title",
    "e.g., KOKOMO-W220AX-IS, ESG620",
    "MMM DD, YYYY",
    "MMM DD, YYYY",
    "MMM DD, YYYY",
    "Resolved / Open",
    "What the customer reported (paragraph)",
    "Bullet points: root cause, resolution, impact (use • for bullets)",
    "Full narrative details (paragraph)",
]
for i, hint in enumerate(hints, 1):
    cell = ws2.cell(row=3, column=i, value=hint)
    cell.font = hint_font
    cell.fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
    cell.alignment = wrap_align
    cell.border = thin_border
ws2.row_dimensions[3].height = 50

# Example data (March 2026 tickets)
category_fills = {
    "T2 Issue": LIGHT_BLUE,
    "Tracker": LIGHT_PURPLE,
    "Customer Conversation": LIGHT_GREEN,
    "Email Conversation": LIGHT_YELLOW,
    "Appendix": LIGHT_GRAY,
}

examples = [
    [
        "T2 Issue", "82514414",
        "Summary of RADIUS Test Failure (KOKOMOCloud / APRESIA)",
        "KOKOMO-W220AX-IS",
        "Mar 12, 2026", "Mar 30, 2026", "Apr 2, 2026", "Resolved",
        "When testing the KOKOMOCloud feature while using Custom RADIUS under WPA3-Enterprise, a certificate missing error is displayed even though RadSec is disabled.",
        "• The Phenomenon: RADIUS test under WPA3-Enterprise showed certificate missing error even with RadSec disabled\n• Root Cause: Protocol mismatch — script hardcoded GTC, customer required MS-CHAPv2\n• Misleading Error Logic: Generic RadSec error message for any RADIUS failure\n• Immediate Resolution: R&D patched AP via SSH to support both MS-CHAPv2 and GTC\n• Long-term: Fix in next firmware release + UI error message update (ETA: Late April 2026)",
        "In March 2026, a significant technical discrepancy was identified within the KOKOMOCloud and APRESIA monitoring platforms concerning the KOKOMO-W220AX-IS Access Point (AP)...",
    ],
    [
        "T2 Issue", "83633969",
        '"Web Authentication Failed" Error',
        "KOKOMOCloud (Web Portal)",
        "Mar 16, 2026", "Mar 19, 2026", "Apr 2, 2026", "Resolved",
        "User so-aoi@iij.ad.jp reported unable to log in from March 11-16 due to 'web authentication failed' error. Issue self-resolved.",
        "• Root Cause: Google reCAPTCHA v3 blocked login due to low risk score\n• Risk Factors: Low IP reputation, VPN/proxy, ad blockers, mechanical browsing patterns\n• Persistent Failure: reCAPTCHA evaluates IP-level, so switching browsers didn't help\n• Resolution: Resolved naturally as IP reputation improved\n• Mitigation: Switch networks (mobile hotspot) or implement whitelist mechanism",
        "Between March 11 and March 16, 2026, a persistent login issue was reported for the user account so-aoi@iij.ad.jp...",
    ],
    [
        "T2 Issue", "86120230",
        "Unable to Use Activity and Internet Connectivity in Diagnostic Tools",
        "KOKOMOCloud (Diagnostic Tools)",
        "Mar 26, 2026", "Mar 31, 2026", "Mar 31, 2026", "Resolved",
        "Activity and Internet Connectivity pages in diagnostic tools fail to load, stuck on loading spinner. Issue only occurs in corporate network with proxy.",
        "• Issue: Diagnostic pages stuck on loading screen\n• Environment: Corporate network with TLS/SSL inspection proxy only\n• Root Cause: Pages rely on WebSocket (WSS) — proxy disrupts the handshake\n• Resolution: Exclude dolphin-jp-*.production.kokomocloud.com from TLS/SSL inspection\n• Outcome: Customer guidance provided, ticket closed",
        "This report outlines an issue encountered when accessing the Activity and Internet Connectivity pages within the diagnostic tools of a cloud-managed device...",
    ],
]

for row_idx, ticket in enumerate(examples, 4):
    cat = ticket[0]
    fill = category_fills.get(cat, None)
    for col_idx, val in enumerate(ticket, 1):
        cell = style_cell(ws2, row_idx, col_idx, fill=fill)
        cell.value = val
    ws2.row_dimensions[row_idx].height = 120

# Column widths
col_widths = [22, 14, 35, 24, 14, 14, 14, 12, 45, 55, 55]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

# Freeze panes
ws2.freeze_panes = "A4"

# Data validation for Category
from openpyxl.worksheet.datavalidation import DataValidation
cat_dv = DataValidation(
    type="list",
    formula1='"T2 Issue,Tracker,Customer Conversation,Email Conversation,Appendix"',
    allow_blank=True,
)
cat_dv.error = "Please select a valid category"
cat_dv.errorTitle = "Invalid Category"
ws2.add_data_validation(cat_dv)
cat_dv.add(f"A4:A100")

status_dv = DataValidation(
    type="list",
    formula1='"Resolved,Open"',
    allow_blank=True,
)
ws2.add_data_validation(status_dv)
status_dv.add(f"H4:H100")


# ═══════════════════════════════════════
# Sheet 3: Instructions
# ═══════════════════════════════════════
ws3 = wb.create_sheet("Instructions")
ws3.sheet_properties.tabColor = "D97706"

ws3.merge_cells("A1:B1")
inst_title = ws3["A1"]
inst_title.value = "How to Use This Template"
inst_title.font = Font(bold=True, color=WHITE, size=14, name="Arial")
inst_title.fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid")
inst_title.alignment = Alignment(horizontal="center")
ws3.row_dimensions[1].height = 36

instructions = [
    ("Step 1", "Fill in the 'Report Info' sheet with the month, author, and ticket counts."),
    ("Step 2", "Go to the 'Tickets' sheet and add each ticket as one row."),
    ("Step 3", "Category column has a dropdown — select from: T2 Issue, Tracker, Customer Conversation, Email Conversation, Appendix."),
    ("Step 4", "For Case Summary, use bullet points with • at the start of each line."),
    ("Step 5", "Affected Product/Model — tag the specific device or platform involved (e.g., KOKOMO-W220AX-IS, ESG620, KOKOMOCloud)."),
    ("Step 6", "Status column has a dropdown — select Resolved or Open."),
    ("Step 7", "When finished, share this file or export as .xlsx. Claude will read it and generate the HTML report automatically."),
    ("", ""),
    ("Note", "For Customer Conversation and Email Conversation categories:"),
    ("", "  • 'Case Description' column → write the 'Highlight' bullet points"),
    ("", "  • 'Case Summary' column → write the 'Summary' paragraph"),
    ("", "  • 'Ticket Details' column → leave empty (not needed for these categories)"),
    ("", ""),
    ("Tip", "The example data in the Tickets sheet is from the March 2026 report. Clear those rows and replace with new data each month."),
]

for idx, (label, text) in enumerate(instructions, 3):
    cell_a = ws3.cell(row=idx, column=1, value=label)
    cell_a.font = label_font if label else value_font
    cell_b = ws3.cell(row=idx, column=2, value=text)
    cell_b.font = value_font
    cell_b.alignment = wrap_align

ws3.column_dimensions["A"].width = 12
ws3.column_dimensions["B"].width = 80

# Save
output_path = "/Users/terrelyeh/Downloads/Temp/FAE Report/FAE_Report_Template.xlsx"
wb.save(output_path)
print(f"Template saved to: {output_path}")
